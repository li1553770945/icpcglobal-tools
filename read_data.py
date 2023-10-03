from openpyxl import load_workbook
from constant import *
from domain import *

def read_excel_data(file_name):
    schools = list()
    school_dict = dict()
    wb = load_workbook(file_name)
    ws = wb.active
    for row in range(2, ws.max_row + 1):
        team_type = ws.cell(row, TEAM_TYPE_COL).value
        if team_type != OFFICIAL_TEAM_NAME:
            continue

        school_name = ws.cell(row, SCHOOL_NAME_COL).value
        if school_name in school_dict:
            school = school_dict[school_name]
        else:
            school_english_name = ws.cell(row, SCHOOL_ENGLISH_NAME_COL).value
            school = School(school_name, school_english_name)
            school_dict[school_name] = school
            schools.append(school)

        coach_name = ws.cell(row, COACH_NAME_COL).value
        coach_english_name = ws.cell(row, COACH_ENGLISH_NAME_COL).value
        coach_email = ws.cell(row, COACH_EMAIL_COL).value
        coach = Person(coach_name, coach_english_name, coach_email)

        contestant1_name = ws.cell(row, CONTESTANT1_NAME_COL).value
        contestant1_english_name = ws.cell(row, CONTESTANT1_ENGLISH_NAME_COL).value
        contestant1_email = ws.cell(row, CONTESTANT1_EMAIL_COL).value
        contestant1 = Person(contestant1_name, contestant1_english_name, contestant1_email)

        contestant2_name = ws.cell(row, CONTESTANT2_NAME_COL).value
        contestant2_english_name = ws.cell(row, CONTESTANT2_ENGLISH_NAME_COL).value
        contestant2_email = ws.cell(row, CONTESTANT2_EMAIL_COL).value
        contestant2 = Person(contestant2_name, contestant2_english_name, contestant2_email)

        contestant3_name = ws.cell(row, CONTESTANT3_NAME_COL).value
        contestant3_english_name = ws.cell(row, CONTESTANT3_ENGLISH_NAME_COL).value
        contestant3_email = ws.cell(row, CONTESTANT3_EMAIL_COL).value
        contestant3 = Person(contestant3_name, contestant3_english_name, contestant3_email)

        contestants = [contestant1, contestant2, contestant3]

        team_name = ws.cell(row, TEAM_NAME_COL).value
        team_english_name = ws.cell(row,TEAM_ENGLISH_NAME_COL).value
        team_id = ws.cell(row, TEAM_ID_COL).value
        team = Team(school_name,team_name,team_english_name,team_id,contestants,coach)

        school.teams.append(team)

    return schools
